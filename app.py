import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import datetime
import io
import re
import html
import numpy as np
import json
from pathlib import Path

# ==============================================================================
# CONFIGURACIÓN DE PÁGINA
# ==============================================================================
st.set_page_config(
    page_title="SOV · Conteo v3",
    page_icon="📡",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ==============================================================================
# CSS — tema claro, tipografía limpia, acentos azul corporativo
# ==============================================================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

/* Fondo general blanco */
.stApp { background: #f8fafc; color: #1e293b; }

/* Sidebar */
section[data-testid="stSidebar"] {
    background: #ffffff !important;
    border-right: 1px solid #e2e8f0 !important;
}

/* Header */
.sov-header {
    background: linear-gradient(135deg, #1e3a5f 0%, #1a4a7a 60%, #1e3a5f 100%);
    border-radius: 14px;
    padding: 28px 36px;
    margin-bottom: 24px;
    position: relative;
    overflow: hidden;
}
.sov-header::after {
    content: '';
    position: absolute;
    bottom: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, transparent, #60a5fa, #3b82f6, transparent);
}
.sov-header h1 { font-size: 1.8rem; font-weight: 700; color: #f0f9ff; margin: 0 0 4px 0; letter-spacing: -0.02em; }
.sov-header p  { color: #93c5fd; font-size: 0.88rem; margin: 0; }
.sov-badge {
    display: inline-block;
    background: #ffffff18;
    border: 1px solid #ffffff30;
    color: #bfdbfe;
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.68rem;
    padding: 2px 10px;
    border-radius: 20px;
    margin-bottom: 8px;
    letter-spacing: 0.1em;
    text-transform: uppercase;
}

/* Tarjetas de métricas */
.metric-card {
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 12px;
    padding: 18px;
    text-align: center;
    box-shadow: 0 1px 4px #1e293b0a;
    transition: box-shadow 0.2s ease;
}
.metric-card:hover { box-shadow: 0 4px 14px #3b82f620; border-color: #bfdbfe; }
.metric-card .metric-value {
    font-family: 'JetBrains Mono', monospace;
    font-size: 1.9rem;
    font-weight: 700;
    color: #1d4ed8;
    display: block;
    line-height: 1;
}
.metric-card .metric-label {
    font-size: 0.72rem;
    color: #64748b;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-top: 6px;
    display: block;
}

/* Tarjeta de archivo */
.file-card {
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 12px;
    padding: 14px 18px;
    margin-bottom: 10px;
    box-shadow: 0 1px 3px #1e293b08;
}
.file-card .file-name {
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.8rem;
    color: #475569;
    margin-bottom: 6px;
}
.file-card .client-tag {
    background: #eff6ff;
    border: 1px solid #bfdbfe;
    color: #1d4ed8;
    font-size: 0.71rem;
    padding: 2px 10px;
    border-radius: 20px;
    font-weight: 500;
}
.file-card .warn-tag {
    background: #fffbeb;
    border: 1px solid #fde68a;
    color: #92400e;
    font-size: 0.71rem;
    padding: 2px 10px;
    border-radius: 20px;
}

/* Bloque de cliente manual */
.manual-client-block {
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-left: 3px solid #3b82f6;
    border-radius: 10px;
    padding: 14px 18px 6px 18px;
    margin-bottom: 10px;
}
.manual-client-block h4 {
    color: #1e40af;
    font-size: 0.8rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    margin: 0 0 10px 0;
}

/* Info box */
.info-box {
    background: #eff6ff;
    border: 1px solid #bfdbfe;
    border-radius: 10px;
    padding: 12px 16px;
    font-size: 0.82rem;
    color: #1e40af;
    margin-bottom: 14px;
}

/* Guía de orden */
.order-guide {
    background: #f8fafc;
    border: 1px solid #e2e8f0;
    border-radius: 12px;
    padding: 0;
    overflow: hidden;
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.78rem;
}
.order-guide-row {
    display: flex;
    align-items: center;
    padding: 7px 14px;
    border-bottom: 1px solid #f1f5f9;
    gap: 10px;
}
.order-guide-row:last-child { border-bottom: none; }
.order-guide-row:nth-child(even) { background: #f8fafc; }
.order-guide-row:nth-child(odd)  { background: #ffffff; }
.order-num  { color: #94a3b8; width: 22px; text-align: right; flex-shrink: 0; font-size: 0.72rem; }
.order-tipo { color: #64748b; width: 210px; flex-shrink: 0; }
.order-val  { font-weight: 700; color: #1d4ed8; min-width: 36px; text-align: right; flex-shrink: 0; }
.order-client { color: #1e293b; flex: 1; font-size: 0.74rem; }
.order-header { background: #1e3a5f !important; color: #e0f2fe; font-weight: 600; }
.order-header .order-num,
.order-header .order-tipo,
.order-header .order-val,
.order-header .order-client { color: #e0f2fe !important; }
.order-highlight { background: #eff6ff !important; }

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    background: #f1f5f9;
    border-radius: 10px;
    padding: 3px;
    border: 1px solid #e2e8f0;
    gap: 2px;
}
.stTabs [data-baseweb="tab"] {
    background: transparent;
    border-radius: 8px;
    color: #64748b;
    font-size: 0.82rem;
    font-weight: 500;
    padding: 7px 16px;
    border: none !important;
}
.stTabs [aria-selected="true"] {
    background: #ffffff !important;
    color: #1d4ed8 !important;
    box-shadow: 0 1px 4px #1e293b14;
}

/* Botones */
.stButton > button[kind="primary"] {
    background: #1d4ed8 !important;
    border: none !important;
    border-radius: 8px !important;
    color: white !important;
    font-weight: 600 !important;
    font-size: 0.84rem !important;
    padding: 9px 22px !important;
}
.stButton > button[kind="primary"]:hover {
    background: #1e40af !important;
    box-shadow: 0 4px 14px #1d4ed840 !important;
}
.stButton > button[kind="secondary"] {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 8px !important;
    color: #475569 !important;
    font-weight: 500 !important;
    font-size: 0.82rem !important;
}

/* Botones de descarga */
.stDownloadButton > button {
    background: #ffffff !important;
    border: 1px solid #bfdbfe !important;
    border-radius: 8px !important;
    color: #1d4ed8 !important;
    font-size: 0.8rem !important;
    font-weight: 500 !important;
    width: 100%;
}
.stDownloadButton > button:hover {
    background: #eff6ff !important;
    border-color: #3b82f6 !important;
}

/* Number inputs */
div[data-testid="stNumberInput"] input {
    background: #f8fafc !important;
    border-color: #e2e8f0 !important;
    color: #1e293b !important;
    font-family: 'JetBrains Mono', monospace !important;
    border-radius: 8px !important;
}

/* Textarea (columna de números) */
textarea {
    background: #f8fafc !important;
    border-color: #e2e8f0 !important;
    color: #1d4ed8 !important;
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.85rem !important;
    line-height: 2 !important;
}

/* DataFrames */
.stDataFrame { border-radius: 10px; overflow: hidden; }

/* File uploader */
div[data-testid="stFileUploader"] {
    background: #ffffff !important;
    border: 1.5px dashed #bfdbfe !important;
    border-radius: 12px !important;
}

/* Separadores */
hr { border-color: #e2e8f0 !important; margin: 20px 0 !important; }

/* Estado vacío */
.empty-state {
    text-align: center;
    padding: 60px 20px;
    color: #94a3b8;
}
.empty-state .icon { font-size: 2.8rem; margin-bottom: 14px; }
.empty-state p { font-size: 0.92rem; color: #64748b; margin: 0; }
.empty-state small { font-size: 0.78rem; color: #94a3b8; margin-top: 6px; display: block; }
</style>
""", unsafe_allow_html=True)

# ==============================================================================
# CONSTANTES
# ==============================================================================

CONTEO_TEMPLATE = [
    ("Chery 01-18 | |19-31",                          "Codificación Audiovisuales", "ANCHERY"),
    ("Chery 01-18 | |19-31",                          "Codificación Impresos",      "ANCHERY"),
    ("Chery 01-18 | |19-31",                          "Notas Audiovisuales",        "ANCHERY"),
    ("Chery 01-18 | |19-31",                          "Notas Impresos",             "ANCHERY"),
    ("Chery - Changan, Competencias",                 "Codificación Audiovisuales", "ANCHERY"),
    ("Chery - Changan, Competencias",                 "Codificación Impresos",      "ANCHERY"),
    ("Chery - Changan, Competencias",                 "Notas Audiovisuales",        "ANCHERY"),
    ("Chery - Changan, Competencias",                 "Notas Impresos",             "ANCHERY"),
    ("Comfenalco Valle",                              "Codificación Audiovisuales", "ACOMFEVALLE"),
    ("Comfenalco Valle",                              "Codificación Impresos",      "ACOMFEVALLE"),
    ("Comfenalco Valle",                              "Notas Audiovisuales",        "ACOMFEVALLE"),
    ("Comfenalco Valle",                              "Notas Impresos",             "ACOMFEVALLE"),
    ("Federación Nacional de Avicultores de Colombia","Codificación Audiovisuales", "ANFENAVI"),
    ("Federación Nacional de Avicultores de Colombia","Codificación Impresos",      "ANFENAVI"),
    ("Federación Nacional de Avicultores de Colombia","Notas Audiovisuales",        "ANFENAVI"),
    ("Federación Nacional de Avicultores de Colombia","Notas Impresos",             "ANFENAVI"),
    ("Fundación Santa Fe de Bogotá",                  "Codificación Audiovisuales", "FSANTAFE_AN"),
    ("Fundación Santa Fe de Bogotá",                  "Codificación Impresos",      "FSANTAFE_AN"),
    ("Fundación Santa Fe de Bogotá",                  "Notas Audiovisuales",        "FSANTAFE_AN"),
    ("Fundación Santa Fe de Bogotá",                  "Notas Impresos",             "FSANTAFE_AN"),
    ("Nissan",                                        "Codificación Audiovisuales", "ANNISSAN"),
    ("Nissan",                                        "Codificación Impresos",      "ANNISSAN"),
    ("Nissan",                                        "Notas Audiovisuales",        "ANNISSAN"),
    ("Nissan",                                        "Notas Impresos",             "ANNISSAN"),
    ("Nissan, Competencia",                           "Codificación Audiovisuales", "ANNISSAN"),
    ("Nissan, Competencia",                           "Codificación Impresos",      "ANNISSAN"),
    ("Nissan, Competencia",                           "Notas Audiovisuales",        "ANNISSAN"),
    ("Nissan, Competencia",                           "Notas Impresos",             "ANNISSAN"),
    ("Tigo",                                          "Codificación Audiovisuales", "TIGOAN"),
    ("Tigo",                                          "Codificación Impresos",      "TIGOAN"),
    ("Tigo",                                          "Notas Audiovisuales",        "TIGOAN"),
    ("Tigo",                                          "Notas Impresos",             "TIGOAN"),
    ("Universidad Simón Bolívar",                     "Codificación Audiovisuales", "USIMONAN"),
    ("Universidad Simón Bolívar",                     "Codificación Impresos",      "USIMONAN"),
    ("Universidad Simón Bolívar",                     "Notas Audiovisuales",        "USIMONAN"),
    ("Universidad Simón Bolívar",                     "Notas Impresos",             "USIMONAN"),
    ("Universidad Tecnológica de Bolívar",            "Codificación Audiovisuales", "UTB_AN"),
    ("Universidad Tecnológica de Bolívar",            "Codificación Impresos",      "UTB_AN"),
    ("Universidad Tecnológica de Bolívar",            "Notas Audiovisuales",        "UTB_AN"),
    ("Universidad Tecnológica de Bolívar",            "Notas Impresos",             "UTB_AN"),
]

UNIQUE_CLIENTS = list(dict.fromkeys(c for c, _, _ in CONTEO_TEMPLATE))

REPLICATED_CLIENTS = {
    "Chery 01-18 | |19-31",
    "Chery - Changan, Competencias",
    "Nissan",
    "Nissan, Competencia",
}

_CODE_TO_CLIENT_SIMPLE = {
    "acomfevalle": "Comfenalco Valle",
    "anfenavi":    "Federación Nacional de Avicultores de Colombia",
    "fsantafe_an": "Fundación Santa Fe de Bogotá",
    "tigoan":      "Tigo",
    "usimonan":    "Universidad Simón Bolívar",
    "utb_an":      "Universidad Tecnológica de Bolívar",
}

_LEGACY_KEYWORDS = {
    "Comfenalco Valle":                               ["comfe", "comfenalco"],
    "Federación Nacional de Avicultores de Colombia": ["fenavi", "avicultores", "avicola"],
    "Fundación Santa Fe de Bogotá":                   ["fsant", "santa", "santafe"],
    "Tigo":                                           ["tigo"],
    "Universidad Simón Bolívar":                      ["simon", "usimon", "usim"],
    "Universidad Tecnológica de Bolívar":             ["utb", "tecnologica"],
}

# ==============================================================================
# FUNCIONES AUXILIARES
# ==============================================================================

def find_config_path():
    base = Path(__file__).parent
    for f in base.iterdir():
        if f.suffix.lower() == '.xlsx' and 'config' in f.stem.lower():
            return f
    return None

CONFIG_PATH = find_config_path()


def extract_link_from_cell(cell):
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    return None


def convert_html_entities(text):
    if not isinstance(text, str):
        return text
    text = html.unescape(text)
    html_entities = {
        '&#xF3;': 'ó', '&#xE1;': 'á', '&#xE9;': 'é', '&#xED;': 'í',
        '&#xFA;': 'ú', '&#xF1;': 'ñ', '&#xDC;': 'Ü', '&#xFC;': 'ü',
        '&#xC1;': 'Á', '&#xC9;': 'É', '&#xCD;': 'Í', '&#xD3;': 'Ó',
        '&#xDA;': 'Ú', '&#xD1;': 'Ñ', '&#xC7;': 'Ç', '&#xE7;': 'ç',
    }
    for entity, char in html_entities.items():
        text = text.replace(entity, char)

    def replace_hex(m):
        try: return chr(int(m.group(1), 16))
        except: return m.group(0)

    def replace_dec(m):
        try: return chr(int(m.group(1)))
        except: return m.group(0)

    text = re.sub(r'&#x([0-9A-Fa-f]+);', replace_hex, text)
    text = re.sub(r'&#(\d+);', replace_dec, text)
    for bad, good in {'\u201c': '"', '\u201d': '"', '\u2018': "'", '\u2019': "'",
                      'Â': '', 'â': '', '€': '', '™': ''}.items():
        text = text.replace(bad, good)
    return text


def clean_text(text):
    if not isinstance(text, str):
        return text
    return convert_html_entities(text).strip()


def clean_cuerpo(text):
    if not isinstance(text, str) or text.strip() == '':
        return text
    text = convert_html_entities(text)
    text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', text)
    return text.strip()


def get_client_category(filename):
    fn = Path(filename).stem.lower()
    fn_clean = re.sub(r'^[\d\s\-_]+', '', fn).strip()
    tokens = [t for t in re.split(r'[^a-z0-9]', fn_clean) if t]
    comp_kws = {"c", "com", "comp", "competencia", "competencias", "changan"}
    is_comp = any(t in comp_kws for t in tokens)

    if "anchery"  in fn_clean: return "Chery - Changan, Competencias" if is_comp else "Chery 01-18 | |19-31"
    if "annissan" in fn_clean: return "Nissan, Competencia" if is_comp else "Nissan"
    for code, client in _CODE_TO_CLIENT_SIMPLE.items():
        if code in fn_clean: return client
    if any(kw in t for t in tokens for kw in ["chery"]):
        return "Chery - Changan, Competencias" if is_comp else "Chery 01-18 | |19-31"
    if any(kw in t for t in tokens for kw in ["niss", "nissan"]):
        return "Nissan, Competencia" if is_comp else "Nissan"
    for client, kws in _LEGACY_KEYWORDS.items():
        if any(kw in t for t in tokens for kw in kws): return client
    return None


def build_consolidated_conteo(processed_files, manual_codif=None):
    manual_codif = manual_codif or {}
    client_av, client_graf = {}, {}

    for item in processed_files:
        mc = get_client_category(item['filename'])
        if mc:
            client_av[mc]   = client_av.get(mc, 0)   + item['av_count']
            client_graf[mc] = client_graf.get(mc, 0) + item['grafica_count']

    rows = []
    for client, tipo, codigo in CONTEO_TEMPLATE:
        av    = client_av.get(client, 0)
        graf  = client_graf.get(client, 0)
        m_av  = int(manual_codif.get(client, {}).get('av', 0) or 0)
        m_imp = int(manual_codif.get(client, {}).get('impresos', 0) or 0)
        is_rep = client in REPLICATED_CLIENTS

        if   tipo == "Notas Audiovisuales":        val = av
        elif tipo == "Notas Impresos":             val = graf
        elif tipo == "Codificación Audiovisuales": val = (av if is_rep else 0) + m_av
        elif tipo == "Codificación Impresos":      val = (graf if is_rep else 0) + m_imp
        else:                                      val = 0

        rows.append({"Cliente / Categoría": client, "Tipo de Conteo": tipo,
                     "Código": codigo, "Cantidad": val})
    return pd.DataFrame(rows)


def to_excel_consolidated(df_conteo):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Conteo Consolidado"
    hf = PatternFill("solid", fgColor="1E3A5F")
    hfont = Font(bold=True, color="FFFFFF", name="Calibri")
    bot  = Border(bottom=Side(style='thin', color="E2E8F0"))
    alt  = PatternFill("solid", fgColor="F8FAFC")

    for ci, cn in enumerate(df_conteo.columns, 1):
        c = ws.cell(row=1, column=ci, value=cn)
        c.font = hfont; c.fill = hf; c.border = bot

    for ri, rd in enumerate(df_conteo.itertuples(index=False), 2):
        for ci, val in enumerate(rd, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            if ri % 2 == 0: c.fill = alt
            c.border = bot

    for col in ws.columns:
        mx = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(mx + 4, 14)

    wb.save(output); output.seek(0)
    return output.getvalue()


def to_excel_from_df(df, final_order, filename, av_count, grafica_count):
    output = io.BytesIO()
    cols = [c for c in final_order if c in df.columns]
    df_out = df[cols].copy()
    for col in df_out.columns:
        if hasattr(df_out[col].dtype, 'pyarrow_dtype'):
            df_out[col] = df_out[col].astype(object)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Resultado'
    hf    = PatternFill("solid", fgColor="1E3A5F")
    hfont = Font(bold=True, color="FFFFFF", name="Calibri")

    for i, cn in enumerate(df_out.columns, 1):
        c = ws.cell(row=1, column=i, value=cn)
        c.font = hfont; c.fill = hf

    link_cols = {'Link Nota', 'Link (Streaming - Imagen)'}
    for ri, rd in enumerate(df_out.itertuples(index=False), 2):
        for ci, value in enumerate(rd, 1):
            cn = df_out.columns[ci - 1]
            cell = ws.cell(row=ri, column=ci)
            if cn == 'Fecha' and pd.notna(value):
                if isinstance(value, pd.Timestamp):
                    cell.value = value.to_pydatetime()
                    cell.number_format = 'DD/MM/YYYY'
                else:
                    cell.value = value
            elif cn in link_cols and pd.notna(value) and isinstance(value, str) and value.startswith('http'):
                cell.value = 'Link'; cell.hyperlink = value
                cell.font = Font(color="1D4ED8", underline="single")
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.value = value if pd.notna(value) else None

    for i, cn in enumerate(df_out.columns, 1):
        ltr = ws.cell(row=1, column=i).column_letter
        if cn in ['Título', 'Resumen - Aclaracion']:         ws.column_dimensions[ltr].width = 50
        elif cn in ['Link Nota', 'Link (Streaming - Imagen)']: ws.column_dimensions[ltr].width = 15
        else:                                                  ws.column_dimensions[ltr].width = 20

    ws2 = wb.create_sheet(title='Conteo')
    df_c = build_consolidated_conteo(
        [{'filename': filename, 'av_count': av_count, 'grafica_count': grafica_count}]
    )
    for ci, cn in enumerate(df_c.columns, 1):
        c = ws2.cell(row=1, column=ci, value=cn)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = hf
    for ri, rd in enumerate(df_c.itertuples(index=False), 2):
        for ci, val in enumerate(rd, 1):
            ws2.cell(row=ri, column=ci, value=val)
    for col in ws2.columns:
        mx = max(len(str(c.value or '')) for c in col)
        ws2.column_dimensions[col[0].column_letter].width = max(mx + 4, 14)

    wb.save(output); output.seek(0)
    return output.getvalue()


def load_config(config_source):
    sheets = pd.read_excel(config_source, sheet_name=None, engine='openpyxl')
    region_map = pd.Series(
        sheets['Regiones'].iloc[:, 1].values,
        index=sheets['Regiones'].iloc[:, 0].astype(str).str.lower().str.strip()
    ).to_dict()
    internet_map = pd.Series(
        sheets['Internet'].iloc[:, 1].values,
        index=sheets['Internet'].iloc[:, 0].astype(str).str.lower().str.strip()
    ).to_dict()
    return region_map, internet_map


def process_dossier(dossier_file, region_map, internet_map):
    wb = load_workbook(dossier_file)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1] if cell.value is not None]
    rows = []
    for row in sheet.iter_rows(min_row=2):
        if all(c.value is None for c in row): continue
        row_data = dict(zip(headers, [c.value for c in row[:len(headers)]]))
        for lc in ['URL Nota AV', 'URL (Streaming - Imagen)', 'URL Nota', 'Link Nota AV', 'Link (Streaming - Imagen)']:
            if lc in headers:
                idx = headers.index(lc)
                if idx < len(row):
                    ext = extract_link_from_cell(row[idx])
                    if ext: row_data[lc] = ext
        rows.append(row_data)

    df = pd.DataFrame(rows)
    tipo_medio_map = {
        'online': 'Internet', 'internet': 'Internet', 'diario': 'Prensa',
        'am': 'Radio', 'fm': 'Radio', 'aire': 'Televisión', 'cable': 'Televisión',
        'revista': 'Revistas', 'revistas': 'Revistas',
    }
    df['Tipo de Medio'] = (
        df['Tipo de Medio'].astype(str).str.lower().str.strip()
        .map(tipo_medio_map).fillna(df['Tipo de Medio'].astype(str).str.strip())
    )

    is_av       = df['Tipo de Medio'].isin(['Radio', 'Televisión'])
    is_grafica  = df['Tipo de Medio'].isin(['Prensa', 'Internet', 'Revistas'])
    is_internet = df['Tipo de Medio'] == 'Internet'

    df.loc[is_internet, 'Medio'] = (
        df.loc[is_internet, 'Medio'].astype(str).str.lower().str.strip()
        .map(internet_map).fillna(df.loc[is_internet, 'Medio'])
    )
    df['Región'] = df['Medio'].astype(str).str.lower().str.strip().map(region_map)

    df['ID Noticia']              = df.get('NoticiaId', pd.Series(dtype=str))
    df['Fecha']                   = pd.to_datetime(df.get('Fecha', pd.Series(dtype=str)), dayfirst=True, errors='coerce').dt.normalize()
    df['Hora']                    = df.get('Hora', pd.Series(dtype=str))
    df['Sección - Programa']      = df.get('Sección - Programa', pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Título']                  = df.get('Título', pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Autor - Conductor']       = df.get('Autor - Conductor', pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Nro. Pagina']             = df.get('Nro. Pagina', pd.Series(dtype=str))
    df['Dimensión']               = df.get('Dimensioncm2', pd.Series(dtype=str))
    df['Duración - Nro. Caracteres'] = df.get('Duración - Nro. Caracteres', pd.Series(dtype=str))

    df.loc[is_av, 'Dimensión']               = df.loc[is_av, 'Duración - Nro. Caracteres']
    df.loc[is_av, 'Duración - Nro. Caracteres'] = 0

    cpe_av      = df.get('CPE', pd.Series([np.nan] * len(df)))
    cpe_grafica = df.get('Valor de Nota', pd.Series([np.nan] * len(df)))
    df['CPE']   = np.where(is_av, cpe_av, np.where(is_grafica, cpe_grafica, np.nan))

    df['Tier']                  = df.get('Tier', pd.Series(dtype=str))
    df['Audiencia']             = df.get('Audiencia', pd.Series(dtype=str))
    df['Tono']                  = df.get('Tono', pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Tema']                  = df.get('Tematica', pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Temas Generales - Tema'] = df.get('Temas Generales - Tema', pd.Series(dtype=str)).astype(str).apply(clean_text)

    cuerpo_cleaned = df.get('CuerpoEs', pd.Series([''] * len(df))).astype(str).apply(clean_cuerpo)

    def fmt_grafica(text):
        if not isinstance(text, str) or not text.strip(): return text
        parrafos = [p.strip() for p in text.split('\n') if p.strip()]
        return '\n\n'.join(parrafos) if len(parrafos) > 1 else text

    df['Resumen - Aclaracion'] = np.where(is_av, cuerpo_cleaned, cuerpo_cleaned.apply(fmt_grafica))

    url_nota_av  = df.get('URL Nota AV', df.get('Link Nota AV', pd.Series([''] * len(df)))).fillna('').astype(str)
    url_stream   = df.get('URL (Streaming - Imagen)', pd.Series([''] * len(df))).fillna('').astype(str)
    link_nota_av = url_nota_av.str.replace(r'\.com\.ar', '.com.co', regex=True)
    df['Link Nota'] = np.where(is_av, link_nota_av, np.where(is_grafica, url_stream, ''))
    df['Link Nota'] = df['Link Nota'].replace('', np.nan)
    df['Link (Streaming - Imagen)'] = df.get('URL Nota', pd.Series([''] * len(df))).fillna('').astype(str).replace('', np.nan)

    m_av   = df.get('Menciones - Empresa', pd.Series([''] * len(df))).fillna('').astype(str).apply(clean_text)
    m_graf = df.get('Empresa rel.',        pd.Series([''] * len(df))).fillna('').astype(str).apply(clean_text)
    df['Menciones - Empresa'] = np.where(is_av, m_av, np.where(is_grafica, m_graf, m_av))

    rows_exp = []
    for _, row in df.iterrows():
        menciones = [m.strip() for m in str(row['Menciones - Empresa']).split(';') if m.strip()]
        if not menciones:
            rows_exp.append(row.to_dict())
        else:
            for m in menciones:
                nr = row.to_dict(); nr['Menciones - Empresa'] = m; rows_exp.append(nr)

    df = pd.DataFrame(rows_exp).reset_index(drop=True)
    is_av_f     = df['Tipo de Medio'].isin(['Radio', 'Televisión'])
    is_grafica_f = df['Tipo de Medio'].isin(['Prensa', 'Internet', 'Revistas'])
    return df, int(is_av_f.sum()), int(is_grafica_f.sum())


# ==============================================================================
# SESSION STATE
# ==============================================================================
def _init():
    for k, v in {
        'uploader_key':  0,
        'resultados':    [],
        'manual_codif':  {c: {'av': 0, 'impresos': 0} for c in UNIQUE_CLIENTS},
        'manual_saved':  False,
    }.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init()

# ==============================================================================
# HEADER
# ==============================================================================
st.markdown("""
<div class="sov-header">
    <div class="sov-badge">📡 Media Intelligence · SOV</div>
    <h1>Conteo v3</h1>
    <p>Procesa dossiers de monitoreo, calcula conteos y gestiona la codificación manual por cliente.</p>
</div>
""", unsafe_allow_html=True)

# ==============================================================================
# SIDEBAR
# ==============================================================================
with st.sidebar:
    st.markdown("### ⚙️ Configuración")
    st.markdown("---")

    config_source = None
    if CONFIG_PATH is not None:
        st.success(f"✅ Config encontrado: `{CONFIG_PATH.name}`")
        config_source = CONFIG_PATH
    else:
        st.warning("No se encontró `Configuracion.xlsx` en el repositorio.")
        config_upload = st.file_uploader("Subir Configuracion.xlsx", type=["xlsx"], key="config_upload")
        if config_upload:
            config_source = config_upload
            st.success(f"✅ {config_upload.name}")

    st.markdown("---")
    st.markdown("### 💾 Codificación manual")
    st.caption("Exporte el JSON para conservar los valores entre sesiones e impórtelo la próxima vez.")

    json_bytes = json.dumps(st.session_state['manual_codif'], ensure_ascii=False, indent=2).encode('utf-8')
    st.download_button(
        "⬇️ Exportar JSON",
        data=json_bytes,
        file_name=f"codif_manual_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.json",
        mime="application/json",
        key="dl_manual_sidebar",
        use_container_width=True,
    )
    archivo_json = st.file_uploader("⬆️ Importar JSON", type=["json"], key="upload_manual_sidebar")
    if archivo_json is not None:
        try:
            data_cargada = json.load(archivo_json)
            st.session_state['manual_codif'] = {
                c: data_cargada.get(c, {'av': 0, 'impresos': 0}) for c in UNIQUE_CLIENTS
            }
            st.session_state['manual_saved'] = True
            st.success("Codificación importada correctamente.")
            st.rerun()
        except Exception as e:
            st.error(f"Error al leer el archivo: {e}")

    st.markdown("---")
    st.markdown("### 📋 Convención de nombres")
    st.caption(
        "`<fecha> <CÓDIGO> [m|com]`\n\n"
        "Ejemplos: `19 ANCHERY m`, `19 ANNISSAN com`, `19 FSANTAFE_AN`\n\n"
        "Códigos válidos: `ANCHERY`, `ANNISSAN`, `ACOMFEVALLE`, `ANFENAVI`, "
        "`FSANTAFE_AN`, `TIGOAN`, `USIMONAN`, `UTB_AN`"
    )
    st.markdown("---")
    if st.button("🗑️ Limpiar todos los resultados", type="secondary", use_container_width=True):
        st.session_state['resultados'] = []
        st.session_state['uploader_key'] += 1
        st.rerun()


# ==============================================================================
# CUERPO PRINCIPAL
# ==============================================================================

st.markdown("#### 📂 Dossiers para procesar")
uploaded_dossiers = st.file_uploader(
    "Cargue uno o varios archivos `.xlsx`",
    type=["xlsx"],
    accept_multiple_files=True,
    key=f"dossiers_{st.session_state['uploader_key']}",
    label_visibility="collapsed",
)

if uploaded_dossiers:
    st.markdown(
        f'<div class="info-box">📎 {len(uploaded_dossiers)} archivo(s) listos: '
        + ', '.join(f'<code>{f.name}</code>' for f in uploaded_dossiers)
        + '</div>',
        unsafe_allow_html=True,
    )

can_run = bool(uploaded_dossiers and config_source)
col_btn, col_hint = st.columns([2, 5])
with col_btn:
    run_clicked = st.button("▶ Procesar", disabled=not can_run, type="primary", use_container_width=True)
with col_hint:
    if not config_source:
        st.caption("⚠️ Cargue `Configuracion.xlsx` en el panel lateral.")
    elif not uploaded_dossiers:
        st.caption("Cargue al menos un dossier para continuar.")

# --- Procesamiento ---
if run_clicked:
    try:
        region_map, internet_map = load_config(config_source)
    except Exception as e:
        st.error(f"Error cargando Configuracion.xlsx: {e}")
        st.stop()

    final_order = [
        "ID Noticia", "Fecha", "Hora", "Medio", "Tipo de Medio",
        "Sección - Programa", "Región", "Título", "Autor - Conductor",
        "Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres",
        "CPE", "Tier", "Audiencia", "Tono", "Tema",
        "Temas Generales - Tema", "Resumen - Aclaracion",
        "Link Nota", "Link (Streaming - Imagen)", "Menciones - Empresa",
    ]
    nuevos = []
    progress_bar = st.progress(0, text="Iniciando proceso...")
    for i, dossier_file in enumerate(uploaded_dossiers):
        progress_bar.progress(i / len(uploaded_dossiers), text=f"Procesando {dossier_file.name}...")
        try:
            df, av_count, grafica_count = process_dossier(dossier_file, region_map, internet_map)
            excel_data = to_excel_from_df(df, final_order, dossier_file.name, av_count, grafica_count)
            matched = get_client_category(dossier_file.name)
            nuevos.append({
                'nombre': dossier_file.name,
                'graficas': grafica_count,
                'av': av_count,
                'total': len(df),
                'excel': excel_data,
                'matched_client': matched,
                'filename': f"SOV_{dossier_file.name.replace('.xlsx','')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            })
        except Exception as e:
            st.error(f"Error en {dossier_file.name}: {e}")
    progress_bar.progress(1.0, text="✅ Proceso completado")
    st.session_state['resultados'].extend(nuevos)
    st.session_state['uploader_key'] += 1
    st.balloons()
    st.rerun()

# ==============================================================================
# RESULTADOS
# ==============================================================================
if st.session_state.get('resultados'):
    resultados = st.session_state['resultados']
    total_av    = sum(r['av']      for r in resultados)
    total_graf  = sum(r['graficas'] for r in resultados)
    total_notas = sum(r['total']   for r in resultados)
    total_arch  = len(resultados)

    st.markdown("---")
    c1, c2, c3, c4 = st.columns(4)
    for col, label, val in [
        (c1, "Archivos procesados", total_arch),
        (c2, "Total de registros",  total_notas),
        (c3, "🎬 Audiovisuales",    total_av),
        (c4, "🗞️ Gráficas",         total_graf),
    ]:
        col.markdown(
            f'<div class="metric-card">'
            f'<span class="metric-value">{val}</span>'
            f'<span class="metric-label">{label}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )
    st.markdown("")

    tab_archivos, tab_consolidado, tab_manual = st.tabs([
        "📋  Archivos",
        "📊  Conteo consolidado",
        "✏️  Codificación manual",
    ])

    # -------------------------------------------------------------------------
    # TAB 1 — Archivos
    # -------------------------------------------------------------------------
    with tab_archivos:
        for r in resultados:
            st.markdown(
                f'<div class="file-card">'
                f'<div class="file-name">📄 {r["nombre"]}</div>'
                + (f'<span class="client-tag">🎯 {r["matched_client"]}</span>'
                   if r.get("matched_client")
                   else '<span class="warn-tag">⚠️ Cliente no detectado</span>')
                + '</div>',
                unsafe_allow_html=True,
            )
            cg, ca, ct, cdl = st.columns([1, 1, 1, 2])
            cg.metric("Gráficas", r['graficas'])
            ca.metric("Audiovisuales", r['av'])
            ct.metric("Total", r['total'])
            cdl.download_button(
                "📥 Descargar Excel",
                data=r['excel'],
                file_name=r['filename'],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_file_{r['nombre']}_{r['filename']}",
                use_container_width=True,
            )
            st.markdown("")

    # -------------------------------------------------------------------------
    # TAB 2 — Conteo consolidado
    # -------------------------------------------------------------------------
    with tab_consolidado:
        listado = [
            {'filename': r['nombre'], 'av_count': r['av'], 'grafica_count': r['graficas']}
            for r in resultados
        ]
        df_cons = build_consolidated_conteo(listado, st.session_state['manual_codif'])

        col_table, col_numeric = st.columns([3, 1])
        with col_table:
            st.markdown("##### Tabla consolidada")
            st.dataframe(
                df_cons.style.apply(
                    lambda col: [
                        'color: #1d4ed8; font-weight: 700' if v > 0 else 'color: #cbd5e1'
                        for v in col
                    ] if col.name == 'Cantidad' else ['' for _ in col],
                    axis=0,
                ),
                use_container_width=True,
                hide_index=True,
                height=600,
            )
            excel_cons = to_excel_consolidated(df_cons)
            st.download_button(
                "📥 Descargar Excel de conteo",
                data=excel_cons,
                file_name=f"Conteo_SOV_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_consolidado",
                use_container_width=True,
            )

        with col_numeric:
            st.markdown("##### 📋 Copiar columna")
            st.caption("Seleccione y copie directamente en su plantilla de Excel.")
            numeros = df_cons["Cantidad"].astype(str).tolist()
            st.text_area(
                "Valores (40 filas, orden estándar):",
                value="\n".join(numeros),
                height=560,
                key="raw_numeric_area",
            )

    # -------------------------------------------------------------------------
    # TAB 3 — Codificación manual
    # -------------------------------------------------------------------------
    with tab_manual:
        # --- Guía de orden ---
        with st.expander("📖 ¿En qué orden debo ingresar la data si ya tengo notas codificadas?", expanded=True):
            st.markdown("""
La tabla de conteo siempre sigue el **orden estándar de 40 filas** que se muestra abajo.
Para cada cliente hay 4 filas en este orden: **Codificación AV → Codificación Impresos → Notas AV → Notas Impresos**.

**Regla clave:**
- **Notas AV / Notas Impresos** → los calcula la app automáticamente al procesar los dossiers.
- **Codificación AV / Codificación Impresos** → los ingresa usted manualmente aquí si ya tiene notas analizadas de períodos anteriores o de fuentes externas al dossier. La app los suma a lo que calcula de los archivos.

**¿Qué poner en Codificación para cada cliente?**
- **Chery, Nissan** (clientes "replicados"): la Codificación se autocompleta igual que las Notas. Solo necesita el valor manual si hay notas extra analizadas antes.
- **Todos los demás** (Comfenalco, Fenavi, Santa Fe, Tigo, Simón Bolívar, UTB): la Codificación **siempre** viene del campo manual — ingrese aquí el total de notas ya codificadas para ese cliente.
""")
            # Tabla visual del orden
            rows_html = ""
            for i, (client, tipo, codigo) in enumerate(CONTEO_TEMPLATE, 1):
                highlight = "order-highlight" if "Codificación" in tipo else ""
                rows_html += (
                    f'<div class="order-guide-row {highlight}">'
                    f'<span class="order-num">{i}</span>'
                    f'<span class="order-tipo">{tipo}</span>'
                    f'<span class="order-client">{client}</span>'
                    f'<span style="color:#94a3b8;font-size:0.7rem;font-family:monospace">{codigo}</span>'
                    f'</div>'
                )
            st.markdown(
                f'<div class="order-guide">'
                f'<div class="order-guide-row order-header">'
                f'<span class="order-num">#</span>'
                f'<span class="order-tipo">Tipo de Conteo</span>'
                f'<span class="order-client">Cliente</span>'
                f'<span style="font-size:0.7rem">Código</span>'
                f'</div>'
                + rows_html +
                f'</div>',
                unsafe_allow_html=True,
            )
            st.caption("Las filas resaltadas en azul claro son las de Codificación — las que usted puede ingresar manualmente abajo.")

        st.markdown("")
        st.markdown(
            '<div class="info-box">'
            '✏️ Ingrese aquí las cantidades de <strong>Codificación Audiovisuales</strong> y '
            '<strong>Codificación Impresos</strong> para notas analizadas fuera de los dossiers. '
            'Estos valores se <strong>suman</strong> al conteo automático y se reflejan '
            'de inmediato en la pestaña <em>Conteo consolidado</em>.'
            '</div>',
            unsafe_allow_html=True,
        )
        st.info(
            "💡 **Persistencia entre sesiones:** los valores duran mientras no cierre la pestaña del navegador. "
            "Para conservarlos, exporte el JSON desde el panel lateral e impórtelo la próxima vez que ingrese a la app."
        )

        with st.form("form_codif_manual", clear_on_submit=False):
            nuevos_valores = {}
            for client in UNIQUE_CLIENTS:
                st.markdown(
                    f'<div class="manual-client-block"><h4>{client}</h4></div>',
                    unsafe_allow_html=True,
                )
                col_av, col_im = st.columns(2)
                actual = st.session_state['manual_codif'].get(client, {'av': 0, 'impresos': 0})
                val_av = col_av.number_input(
                    "🎬 Codificación Audiovisuales",
                    min_value=0, step=1,
                    value=int(actual.get('av', 0) or 0),
                    key=f"mav_{client}",
                )
                val_im = col_im.number_input(
                    "🗞️ Codificación Impresos",
                    min_value=0, step=1,
                    value=int(actual.get('impresos', 0) or 0),
                    key=f"mim_{client}",
                )
                nuevos_valores[client] = {'av': val_av, 'impresos': val_im}
                st.markdown("")

            col_save, col_reset = st.columns([2, 1])
            guardar = col_save.form_submit_button("💾 Guardar valores", type="primary", use_container_width=True)
            reset   = col_reset.form_submit_button("🔄 Restablecer todo", use_container_width=True)

            if guardar:
                st.session_state['manual_codif'] = nuevos_valores
                st.session_state['manual_saved']  = True
                st.success("✅ Valores guardados. El conteo consolidado ya refleja estos cambios.")
                st.rerun()
            if reset:
                st.session_state['manual_codif'] = {c: {'av': 0, 'impresos': 0} for c in UNIQUE_CLIENTS}
                st.session_state['manual_saved']  = False
                st.rerun()

        if st.session_state.get('manual_saved'):
            tot_av  = sum(v['av']       for v in st.session_state['manual_codif'].values())
            tot_imp = sum(v['impresos'] for v in st.session_state['manual_codif'].values())
            if tot_av > 0 or tot_imp > 0:
                st.markdown("---")
                st.markdown("**Totales ingresados manualmente:**")
                ma, mi = st.columns(2)
                ma.metric("🎬 AV manual total",       tot_av)
                mi.metric("🗞️ Impresos manual total", tot_imp)

else:
    # Estado vacío
    st.markdown("---")
    st.markdown("""
<div class="empty-state">
    <div class="icon">📡</div>
    <p>Cargue los dossiers y haga clic en <strong>Procesar</strong> para comenzar.</p>
    <small>También puede gestionar la codificación manual desde el panel lateral antes de procesar archivos.</small>
</div>
""", unsafe_allow_html=True)

    with st.expander("✏️ Ingresar codificación manual sin procesar dossiers", expanded=False):
        st.markdown(
            '<div class="info-box">'
            'Puede ingresar y guardar su codificación manual ahora. '
            'Los valores se sumarán al conteo cuando procese los archivos.'
            '</div>',
            unsafe_allow_html=True,
        )
        with st.form("form_codif_manual_empty", clear_on_submit=False):
            nuevos_valores = {}
            for client in UNIQUE_CLIENTS:
                st.markdown(
                    f'<div class="manual-client-block"><h4>{client}</h4></div>',
                    unsafe_allow_html=True,
                )
                col_av, col_im = st.columns(2)
                actual = st.session_state['manual_codif'].get(client, {'av': 0, 'impresos': 0})
                val_av = col_av.number_input(
                    "🎬 Codificación Audiovisuales",
                    min_value=0, step=1,
                    value=int(actual.get('av', 0) or 0),
                    key=f"empty_mav_{client}",
                )
                val_im = col_im.number_input(
                    "🗞️ Codificación Impresos",
                    min_value=0, step=1,
                    value=int(actual.get('impresos', 0) or 0),
                    key=f"empty_mim_{client}",
                )
                nuevos_valores[client] = {'av': val_av, 'impresos': val_im}
                st.markdown("")

            cs, cr = st.columns([2, 1])
            if cs.form_submit_button("💾 Guardar", type="primary", use_container_width=True):
                st.session_state['manual_codif'] = nuevos_valores
                st.session_state['manual_saved']  = True
                st.success("✅ Guardado. Cargue los dossiers cuando esté listo.")
                st.rerun()
            if cr.form_submit_button("🔄 Restablecer", use_container_width=True):
                st.session_state['manual_codif'] = {c: {'av': 0, 'impresos': 0} for c in UNIQUE_CLIENTS}
                st.session_state['manual_saved']  = False
                st.rerun()
